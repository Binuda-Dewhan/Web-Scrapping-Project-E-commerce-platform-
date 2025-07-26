import os
import json
import pandas as pd
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
import string

from textblob import TextBlob
from openpyxl import load_workbook
from openpyxl.styles import Font
from wordcloud import WordCloud
import matplotlib.pyplot as plt
import pandas as pd
import os
from openpyxl.utils import get_column_letter

# Set up logging
os.makedirs("logs", exist_ok=True)
logging.basicConfig(
    filename="logs/analysis.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

# Constants
RAW_DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "raw")
SUMMARY_EXCEL_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "reports", "product_analysis.xlsx")

def create_review_analysis_sheet(df):
    print("🔍 Running sentiment analysis...")
    logging.info("🔍 Starting review analysis.")

    try:
        # ✅ Ensure the reports folder exists
        os.makedirs("reports", exist_ok=True)

        review_rows = []
        for _, row in df.iterrows():
            name = row.get("name", "")
            brand = row.get("brand", "")
            reviews = row.get("all_reviews", [])
            if isinstance(reviews, list):
                for r in reviews:
                    body = r.get("body", "").strip()
                    if body:
                        try:
                            polarity = TextBlob(body).sentiment.polarity
                        except Exception as e:
                            logging.warning(f"Sentiment analysis failed for review: {body[:30]}... - {e}")
                            continue
                        label = "Positive" if polarity > 0.1 else "Negative" if polarity < -0.1 else "Neutral"
                        review_rows.append({
                            "brand": brand,
                            "product": name,
                            "review": body,
                            "sentiment_score": polarity,
                            "sentiment_label": label
                        })

        reviews_df = pd.DataFrame(review_rows)
        if reviews_df.empty:
            print("⚠️ No valid reviews found.")
            logging.warning("⚠️ No valid reviews found for sentiment analysis.")
            return

        # ✅ Save to Excel sheet
        try:
            wb = load_workbook(SUMMARY_EXCEL_PATH)
            if "Review Analysis" in wb.sheetnames:
                del wb["Review Analysis"]
            ws = wb.create_sheet("Review Analysis")

            for col_idx, col in enumerate(reviews_df.columns, 1):
                ws.cell(row=1, column=col_idx, value=col).font = Font(bold=True)
            for row_idx, row in enumerate(reviews_df.itertuples(index=False), start=2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            wb.save(SUMMARY_EXCEL_PATH)
            print("✅ Review Analysis sheet created.")
            logging.info("✅ Review Analysis sheet saved in Excel.")
        except Exception as e:
            logging.error(f"❌ Failed to write Review Analysis to Excel: {e}")

        # ✅ Generate word clouds
        print("☁️ Generating word clouds...")
        for label, color in [("All", "cool"), ("Positive", "Greens"), ("Negative", "Reds")]:
            try:
                if label == "All":
                    texts = reviews_df["review"]
                else:
                    texts = reviews_df[reviews_df["sentiment_label"] == label]["review"]

                if not texts.empty:
                    wc = WordCloud(width=800, height=400, background_color="white", colormap=color).generate(" ".join(texts))

                    plt.figure(figsize=(10, 5))
                    plt.imshow(wc, interpolation="bilinear")
                    plt.axis("off")
                    plt.title(f"{label} Reviews Word Cloud")
                    plt.tight_layout()
                    path = f"reports/{label.lower()}_wordcloud.png"
                    plt.savefig(path)
                    plt.close()
                    logging.info(f"✅ Saved word cloud: {path}")
                    print(f"✅ Saved word cloud: {path}")
            except Exception as e:
                logging.error(f"❌ Failed to generate word cloud for {label} reviews: {e}")

        # ✅ Sentiment distribution plot
        try:
            print("📊 Generating sentiment score plot...")
            plt.figure(figsize=(8, 4))
            reviews_df["sentiment_score"].hist(bins=20, color="skyblue")
            plt.title("Sentiment Score Distribution")
            plt.xlabel("Sentiment Score")
            plt.ylabel("Review Count")
            plt.tight_layout()
            plt.savefig("reports/sentiment_distribution.png")
            plt.close()
            logging.info("✅ Sentiment distribution chart saved.")
            print("✅ Sentiment distribution chart saved.")
        except Exception as e:
            logging.error(f"❌ Failed to generate sentiment score distribution plot: {e}")

        # ✅ Save raw sentiment data
        try:
            reviews_df.to_csv("reports/review_sentiment_data.csv", index=False)
            logging.info("✅ Review sentiment data saved to CSV.")
        except Exception as e:
            logging.error(f"❌ Failed to save sentiment data CSV: {e}")

    except Exception as e:
        logging.error(f"❌ Unexpected error in create_review_analysis_sheet: {e}")
        print(f"❌ An unexpected error occurred: {e}")


def load_all_product_data():
    """
    Loads all product JSON files into a list of dictionaries.
    """
    all_products = []
    for file in os.listdir(RAW_DATA_DIR):
        if file.endswith(".json"):
            path = os.path.join(RAW_DATA_DIR, file)
            try:
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    all_products.append(data)
            except Exception as e:
                logging.warning(f"Failed to load {file}: {e}")
    logging.info(f"Loaded {len(all_products)} products from JSON.")
    return all_products

def create_product_summary_df(products):
    """
    Transforms list of product dicts into a clean, analysis-ready DataFrame.
    """
    df = pd.DataFrame(products)

    # Clean up numeric fields
    df["price"] = pd.to_numeric(df["price"], errors="coerce")
    df["rating"] = pd.to_numeric(df["rating"], errors="coerce")
    df["review_count"] = pd.to_numeric(df["reviews"], errors="coerce")
    

    # Extract brand from name
    df["brand"] = df["name"].str.split().str[0]

    # Combine reviews
    df["All_reviews"] = df["all_reviews"].apply(
        lambda reviews: " ".join(
            [f"{i+1}. {r.get('body', '').strip()}" for i, r in enumerate(reviews)]
        ) if isinstance(reviews, list) else ""
    )

    # Extract specs
    df["ram"] = df["full_specs"].apply(lambda x: x.get("System Memory (RAM)") if isinstance(x, dict) else None)
    df["storage"] = df["full_specs"].apply(lambda x: x.get("Total Storage Capacity") if isinstance(x, dict) else None)
    df["cpu"] = df["full_specs"].apply(lambda x: x.get("Processor Model") if isinstance(x, dict) else None)
    df["model_number"] = df["full_specs"].apply(lambda x: x.get("Model Number") if isinstance(x, dict) else None)
    df["year"] = df["full_specs"].apply(lambda x: x.get("Year of Release") if isinstance(x, dict) else None)
    
        # Extract specs using column names expected by spec comparison
    df["system_memory_ram"] = df["full_specs"].apply(lambda x: x.get("System Memory (RAM)") if isinstance(x, dict) else None)
    df["total_storage_capacity"] = df["full_specs"].apply(lambda x: x.get("Total Storage Capacity") if isinstance(x, dict) else None)
    df["processor_model"] = df["full_specs"].apply(lambda x: x.get("Processor Model") if isinstance(x, dict) else None)
    df["cpu_boost_clock_frequency"] = df["full_specs"].apply(lambda x: x.get("CPU Boost Clock Frequency") if isinstance(x, dict) else None)
    df["number_of_cpu_cores"] = df["full_specs"].apply(lambda x: x.get("Number of CPU Cores") if isinstance(x, dict) else None)
    df["screen_size"] = df["full_specs"].apply(lambda x: x.get("Screen Size") if isinstance(x, dict) else None)
    df["screen_resolution"] = df["full_specs"].apply(lambda x: x.get("Screen Resolution") if isinstance(x, dict) else None)
    df["refresh_rate"] = df["full_specs"].apply(lambda x: x.get("Refresh Rate") if isinstance(x, dict) else None)
    df["brightness"] = df["full_specs"].apply(lambda x: x.get("Brightness") if isinstance(x, dict) else None)
    df["graphics"] = df["full_specs"].apply(lambda x: x.get("Graphics") if isinstance(x, dict) else None)
    df["gpu_brand"] = df["full_specs"].apply(lambda x: x.get("GPU Brand") if isinstance(x, dict) else None)
    df["battery_life_up_to"] = df["full_specs"].apply(lambda x: x.get("Battery Life (up to)") if isinstance(x, dict) else None)
    df["product_weight"] = df["full_specs"].apply(lambda x: x.get("Product Weight") if isinstance(x, dict) else None)
    df["year_of_release"] = df["full_specs"].apply(lambda x: x.get("Year of Release") if isinstance(x, dict) else None)


    # Drop bulky fields
    df.drop(columns=["specs", "full_specs", "reviews"], errors="ignore", inplace=True)

    # Reorder columns
    cols = df.columns.tolist()
    if "brand" in cols:
        cols.insert(0, cols.pop(cols.index("brand")))
        df = df[cols]

    return df

def save_summary_to_excel(df):
    os.makedirs(os.path.dirname(SUMMARY_EXCEL_PATH), exist_ok=True)

    # Save Product Summary
    df.to_excel(SUMMARY_EXCEL_PATH, sheet_name="Product Summary", index=False)

    # Load and format with openpyxl
    wb = load_workbook(SUMMARY_EXCEL_PATH)
    ws = wb["Product Summary"]
    
    # ✅ Format as Excel Table
    last_col_letter = get_column_letter(ws.max_column)
    table_ref = f"A1:{last_col_letter}{ws.max_row}"
    table = Table(displayName="ProductSummaryTable", ref=table_ref)

    # ✅ Format as Excel Table
    # table_ref = f"A1:{string.ascii_uppercase[ws.max_column - 1]}{ws.max_row}"
    # table = Table(displayName="ProductSummaryTable", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

    # ✅ Conditional formatting for price > 900
    price_col = None
    brand_col = None
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header == "price":
            price_col = col
        elif header == "brand":
            brand_col = col

    if price_col:
        col_letter = string.ascii_uppercase[price_col - 1]
        price_range = f"{col_letter}2:{col_letter}{ws.max_row}"

        # Red fill for prices > 900
        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        rule_red = CellIsRule(operator="greaterThan", formula=["900"], fill=red_fill)
        ws.conditional_formatting.add(price_range, rule_red)

        # Green fill for prices <= 900
        green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        rule_green = CellIsRule(operator="lessThanOrEqual", formula=["900"], fill=green_fill)
        ws.conditional_formatting.add(price_range, rule_green)

    # ✅ Data validation dropdown for brand
    if brand_col:
        unique_brands = list(df["brand"].dropna().unique())
        if unique_brands:
            brand_list = ",".join(unique_brands)
            col_letter = string.ascii_uppercase[brand_col - 1]
            dv_range = f"{col_letter}2:{col_letter}{ws.max_row}"
            dv = DataValidation(type="list", formula1=f'"{brand_list}"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(dv_range)

    wb.save(SUMMARY_EXCEL_PATH)
    logging.info(f"✅ Final Excel saved at {SUMMARY_EXCEL_PATH}")
    print(f"✅ Excel file created: {SUMMARY_EXCEL_PATH}")
    
def create_spec_comparison_sheet(df):
    import re
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font
    from openpyxl.worksheet.table import Table, TableStyleInfo
    import string

    # --- Reload workbook ---
    wb = load_workbook(SUMMARY_EXCEL_PATH)
    if "Specifications Comparison" in wb.sheetnames:
        del wb["Specifications Comparison"]
    ws = wb.create_sheet("Specifications Comparison")

    # --- Spec columns to compare ---
    spec_columns = [
        "brand", "name", "price", "rating", "system_memory_ram", "total_storage_capacity",
        "processor_model", "cpu_boost_clock_frequency", "number_of_cpu_cores",
        "screen_size", "screen_resolution", "refresh_rate", "brightness",
        "graphics", "gpu_brand", "battery_life_up_to", "product_weight", "year_of_release"
    ]
    spec_columns = [col for col in spec_columns if col in df.columns]

    # --- Sort by price, rating, brand (if available) ---
    sort_keys = [key for key in ["price", "rating", "brand"] if key in df.columns]
    df = df.sort_values(by=sort_keys, ascending=[True, False, True])

    # --- Write headers ---
    for col_idx, col_name in enumerate(spec_columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name).font = Font(bold=True)

    # --- Write data ---
    for row_idx, row in enumerate(df[spec_columns].itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # --- Yellow fill for columns with unique values ---
    for col_idx, col_name in enumerate(spec_columns, 1):
        values = df[col_name].dropna().unique()
        if len(values) > 1:
            col_letter = string.ascii_uppercase[col_idx - 1]
            yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
            for row in range(2, df.shape[0] + 2):
                ws[f"{col_letter}{row}"].fill = yellow_fill

    # --- Helper to extract first float number from a string ---
    def extract_number(s):
        match = re.search(r"([0-9.]+)", str(s))
        return float(match.group(1)) if match else None

    # --- Highlight best-in-class ---
    def highlight_best(col_name, is_max=True, fill_color="CCFFCC"):
        if col_name not in df.columns:
            return

        numeric_values = df[col_name].apply(extract_number)
        if numeric_values.isnull().all():
            return

        best_value = numeric_values.max() if is_max else numeric_values.min()
        col_idx = spec_columns.index(col_name) + 1
        col_letter = string.ascii_uppercase[col_idx - 1]

        for i, val in enumerate(numeric_values, start=2):
            if val == best_value:
                ws[f"{col_letter}{i}"].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    # ✅ Highlight: Max RAM, Max Battery, Min Weight
    highlight_best("system_memory_ram", is_max=True, fill_color="C6EFCE")       # Green
    highlight_best("battery_life_up_to", is_max=True, fill_color="C6EFCE")      # Green
    highlight_best("product_weight", is_max=False, fill_color="ADD8E6")         # Blue

    # --- Add Excel Table ---
    last_col = string.ascii_uppercase[len(spec_columns) - 1]
    table_range = f"A1:{last_col}{df.shape[0] + 1}"
    table = Table(displayName="SpecComparisonTable", ref=table_range)
    style = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    wb.save(SUMMARY_EXCEL_PATH)
    logging.info("✅ Created 'Specifications Comparison' sheet with highlights.")
    print("✅ Specifications Comparison sheet created with table and highlights.")




if __name__ == "__main__":
    products = load_all_product_data()
    df_summary = create_product_summary_df(products)
    print(df_summary.head(10))
    logging.info("✅ Product summary DataFrame created.")
    
    save_summary_to_excel(df_summary)
    create_spec_comparison_sheet(df_summary)
    create_review_analysis_sheet(df_summary)
